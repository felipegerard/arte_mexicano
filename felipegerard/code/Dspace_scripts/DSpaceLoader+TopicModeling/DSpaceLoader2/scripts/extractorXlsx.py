# -*- coding: utf-8 -*-
"""
#DSpaceLoader V2
/scripts/extractorXlsx.py
#########
#	31/08/2015
#	Sistema desarrollado por el GIL, Instituto de Ingenieria UNAM
#	cgonzalezg@iingen.unam.mx
#	Obtiene los registros del xlsx indicado
#	return: una lista de diccionarios, cada diccionario tiene como llave el nombre de la columna del xlsx
#########
"""
import sys
from openpyxl import load_workbook
import logging

def abrirExcel(direccionArchivo):
	wb = load_workbook(direccionArchivo)
	return wb.active

def extraerLibros(direccionArchivo, descriptor):

	libros = list()

	ws = descriptor.rows
	for fila in range(1,len(ws)):
		if not ws[fila][0].value:
			break
		tmp = dict()
		for columna in range(len(ws[0])):
			tmp[ws[0][columna].value] = ws[fila][columna].value
		libros.append(tmp)
	logging.info("se han extraido "+str(fila)+" registros de "+direccionArchivo)
	print "se han extraido "+str(fila)+" registros de "+direccionArchivo
	return libros

def main(args):	
	if args[0] != "":
		print "NO ES POSIBLE EJECUTAR EL SCRIPT. FAVOR DE UTILIZAR DSpaceLoader.sh"
		return False

	direccionArchivo = args[1]
	print "Abriendo",direccionArchivo

	descriptor = abrirExcel(direccionArchivo)
	libros = extraerLibros(direccionArchivo, descriptor)

	return libros

if __name__ == '__main__':
    main(sys.argv)