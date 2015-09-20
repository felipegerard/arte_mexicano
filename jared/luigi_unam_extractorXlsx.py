# -*- coding: utf-8 -*-
"""
Proceso de extractorPDFsFileSystem de la unam, junta los pdfs
de cada libro en bloques y los une, en una carpeta que se define
por el usuario se guardan estos bloques, tiene una opcion para
no procesar todos los libros sino unos cuantos
"""

import sys
from openpyxl import load_workbook
import logging

class DireccionArchivo(luigi.ExternalTask):
    """
    Vamos a dar de alta el directorio del archivo
    """
    filename = luigi.Parameter()

    def output(self):
        """
        Solo va a regresar el directorio del arcivho
        """

        return luigi.LocalTarget(self.filename)

class ExractorXlsx(luigi.Task):
  input_dir = luigi.Parameter()

  def requires(self):
    return DireccionArchivo(self.input_dir)

  def run(self):

    def abrirExcel(direccionArchivo):
    	wb = load_workbook(direccionArchivo)
		return wb.active

	def extraerLibros(direccionArchivo, descriptor):
		libros = list()
		ws = descriptor.rows
		for fila in range(1,len(ws)):
			if not ws[fila][0].value:
				break
			tmp = dict()
			for columna in range(len(ws[0])):
				tmp[ws[0][columna].value] = ws[fila][columna].value
			libros.append(tmp)
		logging.info("se han extraido "+str(fila)+" registros de "+direccionArchivo)
		print "se han extraido "+str(fila)+" registros de "+direccionArchivo
		return libros

	print "Abriendo", self.input_dir

	descriptor = abrirExcel(self.input_dir)
	libros = extraerLibros(self.input_dir, descriptor)

	f = self.output().open('w')
    pickle.dump(libros, f)
    f.close()

	def output(self):
    	return luigi.LocalTarget('libros.pickle')

if __name__ == '__main__':
    luigi.run()     	        